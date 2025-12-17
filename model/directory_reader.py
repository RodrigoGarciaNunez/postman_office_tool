import openpyxl


class directory_reader:
    
    def __init__(self):
        self.direcoty = dict()
        self.direcoty_format = ['ID', 'NAME', 'TELEFONO', 'EMAIL']

    def leer_directorio(self, file):

       
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            print("voy bien")
            # primero, corroborar el formato
            if len(wb.sheetnames) > 1:
                print("Formato de directorio no valido")
                return
     
            for hoja in wb.sheetnames:
                ws = wb[hoja]

                for i in range(len(self.direcoty_format)):
            

                    if self.direcoty_format[i] is ws.cell(1, i+1).value:
                        print("Formato equivocado")
                        return
                
                for fila in ws.iter_rows(min_row= 2, min_col=2):
                    for celda in fila:
                        id_cell_value = ws.cell(celda.row, 1).value
                        self.direcoty.setdefault(id_cell_value, []).append(celda.value) 
    
                        #print(f"{celda.value}{celda.coordinate}")
                        #self.direcoty[ws.cell(celda.row, 2).value] = celda.value
            
            print(self.direcoty)

            
        except Exception as e:
            print(e)
            print(f"[XLSX] No se pudo leer {file}")
            #self.output_text_2.insert(tk.END, f"[XLSX] No se pudo leer {ruta}")


if __name__ == "__main__":
    directory = dict()
    dr = directory_reader()
    file = "test_resourses/test_dir.xlsx"
    dr.leer_directorio(file)